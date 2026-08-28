# -*- coding: utf-8 -*-
"""
ETF/종목 리스트 자동 업데이트 Streamlit 앱  (pykrx + KIS API 버전)

- 1번 시트: 국내 상장 ETF 전체 (pykrx + KIS API 시가총액)
- 2번 시트: KOSPI200 종목 (pykrx)
- 4번 시트: 미국 S&P500 종목 (위키피디아 + yfinance)
- 5번 시트: 미국 나스닥100 종목 (위키피디아 + yfinance)
- 6번 시트: 미국 배당 ETF(SCHD) 구성종목 (SEC N-PORT + OpenFIGI + yfinance)
- 3번 시트(미국 상장 ETF 전체)는 무료 소스 한계로 이번 버전에서 자동화하지 않음

핵심 설계:
- 국내(pykrx) 파트는 "종목별 반복 호출"을 쓰지 않고,
  기준일 여러 개(오늘/1개월전/6개월전/YTD시작/1년전 영업일)의 "전종목 스냅샷"을
  각각 1회씩만 호출한 뒤 종목코드로 매칭 → 훨씬 빠르고 서버 부담이 적음
- 국내 ETF 시가총액은 pykrx가 제공하지 않아(2026-08-28 확인) KIS Open API로 별도 조회
"""

import os
import io
import time
import json
import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ETF/종목 리스트 업데이트", layout="wide")

# ----------------------------------------------------------------------------
# KRX 로그인 (2026-01-27부터 KRX 정보데이터시스템이 회원제로 전환되어
# pykrx의 일부 함수는 KRX_ID / KRX_PW 환경변수가 필요합니다.)
#
# 우선순위: .env 파일(로컬 개발/테스트용) → 없으면 화면에서 직접 입력받음
# .env 방식만 쓰고 싶다면 USE_ENV_FILE=True로 고정하면 됩니다.
# 화면 입력 방식만 쓰고 싶다면(더 안전) USE_ENV_FILE=False로 바꾸면 됩니다.
# ----------------------------------------------------------------------------
USE_ENV_FILE = True  # 테스트 기간: True(.env 자동 로드) / 배포·공유 시: False(매번 화면 입력)


def ensure_krx_login() -> bool:
    """KRX_ID / KRX_PW 환경변수를 설정. 성공 여부를 반환."""
    if USE_ENV_FILE:
        try:
            from dotenv import load_dotenv
            load_dotenv()
        except ImportError:
            pass

        if os.environ.get("KRX_ID") and os.environ.get("KRX_PW"):
            return True

    with st.sidebar:
        st.subheader("🔐 KRX 로그인")
        st.caption("2026-01-27부터 KRX 정보데이터시스템 데이터 조회에 로그인이 필요합니다. (무료 회원가입)")
        krx_id = st.text_input("KRX 아이디", value=os.environ.get("KRX_ID", ""))
        krx_pw = st.text_input("KRX 비밀번호", type="password", value=os.environ.get("KRX_PW", ""))

    if krx_id and krx_pw:
        os.environ["KRX_ID"] = krx_id
        os.environ["KRX_PW"] = krx_pw
        return True

    return False


# ----------------------------------------------------------------------------
# KIS(한국투자증권) Open API 인증
#
# 국내 ETF 시가총액 조회에 사용(pykrx는 ETF 시가총액을 제공하지 않음, Colab 검증 완료).
# KIS 정책상 "접근토큰은 1일 1회 발급 원칙"이라, 발급받은 토큰을 파일로 캐시해
# 24시간 이내에는 재발급 없이 재사용한다. (2026-08-28 KIS 안내 반영)
# ----------------------------------------------------------------------------
_KIS_TOKEN_CACHE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), ".kis_token_cache.json"
)


def get_kis_access_token(app_key: str, app_secret: str) -> str:
    """KIS Open API 접근토큰 신규 발급 (실제 발급 요청 - 하루 1회만 호출되어야 함)"""
    import requests

    url = "https://openapi.koreainvestment.com:9443/oauth2/tokenP"
    headers = {"content-type": "application/json"}
    body = {
        "grant_type": "client_credentials",
        "appkey": app_key,
        "appsecret": app_secret,
    }
    resp = requests.post(url, headers=headers, json=body, timeout=10)
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_kis_access_token_cached(app_key: str, app_secret: str) -> str:
    """
    KIS 접근토큰을 파일로 캐시해 재사용한다.
    23시간 이내 발급된 토큰이면 재사용하고, 그 이상 지났으면 새로 발급한다.
    """
    now = time.time()

    if os.path.exists(_KIS_TOKEN_CACHE_PATH):
        try:
            with open(_KIS_TOKEN_CACHE_PATH, "r") as f:
                cache = json.load(f)
            if cache.get("app_key") == app_key and now - cache.get("issued_at", 0) < 23 * 3600:
                return cache["access_token"]
        except Exception:
            pass

    token = get_kis_access_token(app_key, app_secret)
    try:
        with open(_KIS_TOKEN_CACHE_PATH, "w") as f:
            json.dump({"access_token": token, "issued_at": now, "app_key": app_key}, f)
    except Exception:
        pass  # 캐시 저장 실패해도 토큰 자체는 정상 반환

    return token


def ensure_kis_auth():
    """
    KIS_APP_KEY / KIS_APP_SECRET을 읽어 접근토큰을 준비한다.
    없으면 사이드바에서 입력받고, 그래도 없으면 None을 반환한다
    (이 경우 국내 ETF 시가총액은 결측으로 처리되고 나머지 기능은 정상 작동).
    """
    if USE_ENV_FILE:
        try:
            from dotenv import load_dotenv
            load_dotenv()
        except ImportError:
            pass

    app_key = os.environ.get("KIS_APP_KEY")
    app_secret = os.environ.get("KIS_APP_SECRET")

    if not (app_key and app_secret):
        with st.sidebar:
            st.subheader("🔑 KIS(한국투자증권) API")
            st.caption("국내 ETF 시가총액 조회에 사용됩니다. 입력하지 않아도 나머지 항목은 정상 작동합니다.")
            app_key = st.text_input("KIS App Key", value=app_key or "", type="password")
            app_secret = st.text_input("KIS App Secret", value=app_secret or "", type="password")

    if not (app_key and app_secret):
        return None

    try:
        access_token = get_kis_access_token_cached(app_key, app_secret)
    except Exception as e:
        st.sidebar.error(f"KIS 토큰 발급 실패: {e}")
        return None

    return {"access_token": access_token, "app_key": app_key, "app_secret": app_secret}


class _KisRateLimiter:
    """KIS 초당 호출 제한(개인 10건/초) 안에서 여러 요청을 병렬로 보내기 위한
    간단한 토큰 버킷 방식 속도 제한기. 여러 스레드가 공유해서 사용한다."""
    def __init__(self, max_per_second: int):
        import threading
        self.max_per_second = max_per_second
        self.lock = threading.Lock()
        self.timestamps = []

    def wait(self):
        while True:
            with self.lock:
                now = time.time()
                self.timestamps = [t for t in self.timestamps if now - t < 1.0]
                if len(self.timestamps) < self.max_per_second:
                    self.timestamps.append(now)
                    return
                sleep_time = 1.0 - (now - self.timestamps[0])
            time.sleep(max(sleep_time, 0.01))


def fetch_kis_market_caps(access_token: str, app_key: str, app_secret: str,
                           tickers: list, max_workers: int = 8, max_per_second: int = 8,
                           progress_cb=None) -> dict:
    """
    KIS Open API로 전 종목 시가총액(억원)을 병렬로 조회한다.
    '국내주식 현재가 시세' API(FHKST01010100)의 hts_avls 필드가
    이미 억원 단위로 계산된 시가총액이라 별도 계산 불필요 (Colab 검증 완료).

    순차 호출은 개별 요청의 네트워크 왕복 시간(0.5~1초)이 그대로 누적되어
    1000종목 이상에서 20분 이상 걸리는 문제가 있어(2026-08-28 로컬 실행에서 확인),
    ThreadPoolExecutor + 속도 제한기로 동시에 여러 요청을 보내 단축한다.
    max_per_second는 KIS 개인 계좌 제한(초당 10건)보다 낮게 잡아 안전하게 운영한다.
    """
    import requests
    from concurrent.futures import ThreadPoolExecutor, as_completed

    url = "https://openapi.koreainvestment.com:9443/uapi/domestic-stock/v1/quotations/inquire-price"
    headers = {
        "content-type": "application/json",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHKST01010100",
    }

    limiter = _KisRateLimiter(max_per_second)
    market_caps = {}
    total = len(tickers)
    done_count = [0]

    def fetch_one(ticker):
        limiter.wait()
        params = {"FID_COND_MRKT_DIV_CODE": "J", "FID_INPUT_ISCD": ticker}
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=5)
            resp.raise_for_status()
            data = resp.json()
            avls = data.get("output", {}).get("hts_avls")
            return ticker, (int(avls) if avls not in (None, "") else None)
        except Exception:
            return ticker, None

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(fetch_one, t): t for t in tickers}
        for future in as_completed(futures):
            ticker, cap = future.result()
            market_caps[ticker] = cap
            done_count[0] += 1
            if progress_cb and done_count[0] % 50 == 0:
                progress_cb(min(0.75 + 0.20 * done_count[0] / total, 0.95),
                            f"시가총액 조회 중... {done_count[0]}/{total}")

    return market_caps


# ----------------------------------------------------------------------------
# 공통 컬럼 정의
# ----------------------------------------------------------------------------
ETF_COLS = [
    "종목코드", "ETF명", "시가총액(억원)", "PR\n1개월", "PR\n6개월", "PR\nYTD", "PR\n1년",
    "최근 1년\n분배율",
]

STOCK_COLS_KR = [
    "종목코드", "종목명", "시가총액(억원)", "현재가", "PR\n1개월", "PR\n6개월", "PR\nYTD", "PR\n1년",
    "최근 1년\n배당수익률",
]

STOCK_COLS_US = [
    "종목코드", "종목명", "시가총액(Mil)", "현재가", "PR\n1개월", "PR\n6개월", "PR\nYTD", "PR\n1년",
    "최근 1년\n배당수익률",
]


# ----------------------------------------------------------------------------
# 공통 유틸: 최근 영업일 구하기
# ----------------------------------------------------------------------------
def _has_valid_close(df: pd.DataFrame) -> bool:
    """스냅샷이 비어있지 않더라도, 휴장일(공휴일)에는 종목은 있지만
    시가/고가/저가/종가가 전부 0으로 채워진 DataFrame이 반환될 수 있다.
    실제 거래된 유효한 종가가 하나라도 있는지 확인한다."""
    if df is None or df.empty or "종가" not in df.columns:
        return False
    return (df["종가"] > 0).any()


def nearest_trading_day_close(get_ohlcv_by_ticker_func, target_date: dt.date, max_back: int = 10, retries: int = 2):
    """target_date부터 최대 max_back일 전까지 뒤로 훑으며
    실제 거래가 있었던(종가가 0이 아닌) 전종목 스냅샷과 그 날짜를 반환."""
    d = target_date
    for _ in range(max_back):
        date_str = d.strftime("%Y%m%d")
        df = None
        for attempt in range(retries + 1):
            try:
                df = get_ohlcv_by_ticker_func(date_str)
            except Exception:
                df = None
            if _has_valid_close(df):
                return df, date_str
            if attempt < retries:
                time.sleep(0.5)
        d -= dt.timedelta(days=1)
    return pd.DataFrame(), None


# ----------------------------------------------------------------------------
# pykrx ETF 종목명 안전 조회
#
# 원인: 2026-05-27 상장된 단일종목 레버리지/인버스 ETF·ETN 18개 종목이
# KRX 원본 데이터 오류로 ELW 목록에도 중복 등재되어 있음(Colab에서 실제 KRX API
# 응답으로 확인, pykrx 원본 함수가 아니라 KRX 데이터 자체의 결함).
# pykrx의 EtxTicker.get_name()은 중복 인덱스일 때 str이 아닌 pandas.Series를
# 반환해 str() 변환 시 지저분한 다중행 텍스트가 나옴.
# ----------------------------------------------------------------------------
def get_etf_ticker_name_safe(ticker: str) -> str:
    from pykrx.website.krx.etx.ticker import EtxTicker

    df = EtxTicker().df

    if ticker not in df.index:
        return None

    rows = df.loc[[ticker]]

    if len(rows) == 1:
        return str(rows.iloc[0]["종목명"])

    # 중복인 경우: ETF/ETN 카테고리 우선 (KRX 데이터 중복 등재 시 ELW가 아닌 쪽 선택)
    preferred = rows[rows["시장"].isin(["ETF", "ETN"])]
    if len(preferred) > 0:
        return str(preferred.iloc[0]["종목명"])

    return str(rows.iloc[0]["종목명"])


# ----------------------------------------------------------------------------
# 1번 시트: 국내 상장 ETF 전체
# ----------------------------------------------------------------------------
def build_domestic_etf_sheet(base_date: dt.date, progress_cb=None, kis_auth=None) -> pd.DataFrame:
    from pykrx import stock

    def snap(date_obj):
        return nearest_trading_day_close(stock.get_etf_ohlcv_by_ticker, date_obj)

    if progress_cb:
        progress_cb(0.05, "기준일 시세 조회 중...")
    today_df, today_str = snap(base_date)
    if today_df.empty:
        raise RuntimeError("기준일 근처에 ETF 시세 데이터가 없습니다. 날짜를 확인해 주세요.")

    if progress_cb:
        progress_cb(0.15, "1개월 전 시세 조회 중...")
    m1_df, _ = snap(base_date - dt.timedelta(days=30))

    if progress_cb:
        progress_cb(0.30, "6개월 전 시세 조회 중...")
    m6_df, _ = snap(base_date - dt.timedelta(days=182))

    if progress_cb:
        progress_cb(0.45, "1년 전 시세 조회 중...")
    y1_df, _ = snap(base_date - dt.timedelta(days=365))

    if progress_cb:
        progress_cb(0.55, "연초(YTD) 시세 조회 중...")
    ytd_df, _ = snap(dt.date(base_date.year, 1, 1))

    if progress_cb:
        progress_cb(0.65, "종목명 매핑 중...")
    tickers = stock.get_etf_ticker_list(today_str)
    name_map = {}
    for code in tickers:
        try:
            name_map[code] = get_etf_ticker_name_safe(code)
        except Exception:
            name_map[code] = None

    market_cap_map = {}
    if kis_auth:
        if progress_cb:
            progress_cb(0.75, f"시가총액 조회 중 (KIS API, {len(tickers)}종목, 수 분 소요)...")
        try:
            market_cap_map = fetch_kis_market_caps(
                kis_auth["access_token"], kis_auth["app_key"], kis_auth["app_secret"],
                tickers, progress_cb=progress_cb,
            )
        except Exception:
            market_cap_map = {}  # 실패 시 시가총액은 결측 처리, 나머지 컬럼은 정상 진행

    def ret(df_ref, code, close_now):
        if df_ref is None or df_ref.empty or code not in df_ref.index:
            return None
        base_price = df_ref.loc[code, "종가"]
        if not base_price:
            return None
        return round(close_now / base_price - 1, 4)

    rows = []
    for code, row in today_df.iterrows():
        close_now = row["종가"]
        rows.append({
            "종목코드": code,
            "ETF명": name_map.get(code),
            "시가총액(억원)": market_cap_map.get(code),
            "PR\n1개월": ret(m1_df, code, close_now),
            "PR\n6개월": ret(m6_df, code, close_now),
            "PR\nYTD": ret(ytd_df, code, close_now),
            "PR\n1년": ret(y1_df, code, close_now),
            "최근 1년\n분배율": None,  # 소스 미확보, 보류 (KIS API에도 해당 필드 없음, 2026-08-28 확인)
        })

    if progress_cb:
        progress_cb(1.0, "완료")
    return pd.DataFrame(rows)[ETF_COLS]


# ----------------------------------------------------------------------------
# 2번 시트: KOSPI200 종목
# ----------------------------------------------------------------------------
def build_kospi200_sheet(base_date: dt.date, progress_cb=None) -> pd.DataFrame:
    from pykrx import stock

    base_date_str = base_date.strftime("%Y%m%d")

    if progress_cb:
        progress_cb(0.05, "KOSPI200 구성종목 조회 중...")
    codes = stock.get_index_portfolio_deposit_file("1028", base_date_str)  # 1028 = 코스피 200

    def snap(date_obj):
        return nearest_trading_day_close(
            lambda d: stock.get_market_ohlcv_by_ticker(d, market="ALL"), date_obj
        )

    if progress_cb:
        progress_cb(0.15, "기준일 시세 조회 중...")
    today_df, today_str = snap(base_date)
    if today_df.empty:
        raise RuntimeError("기준일 근처에 주식 시세 데이터가 없습니다. 날짜를 확인해 주세요.")

    if progress_cb:
        progress_cb(0.25, "1개월 전 시세 조회 중...")
    m1_df, _ = snap(base_date - dt.timedelta(days=30))

    if progress_cb:
        progress_cb(0.40, "6개월 전 시세 조회 중...")
    m6_df, _ = snap(base_date - dt.timedelta(days=182))

    if progress_cb:
        progress_cb(0.55, "1년 전 시세 조회 중...")
    y1_df, _ = snap(base_date - dt.timedelta(days=365))

    if progress_cb:
        progress_cb(0.65, "연초(YTD) 시세 조회 중...")
    ytd_df, _ = snap(dt.date(base_date.year, 1, 1))

    if progress_cb:
        progress_cb(0.75, "배당수익률(DIV) 조회 중...")
    try:
        fund_df = stock.get_market_fundamental(today_str, market="ALL")
    except Exception:
        fund_df = pd.DataFrame()

    if progress_cb:
        progress_cb(0.85, "종목명 매핑 중...")
    name_map = {}
    for code in codes:
        try:
            name_map[code] = stock.get_market_ticker_name(code)
        except Exception:
            name_map[code] = None

    if progress_cb:
        progress_cb(0.90, "시가총액 조회 중...")
    try:
        cap_df = stock.get_market_cap(today_str, market="ALL")
    except Exception:
        cap_df = pd.DataFrame()

    def ret(df_ref, code, close_now):
        if df_ref is None or df_ref.empty or code not in df_ref.index:
            return None
        base_price = df_ref.loc[code, "종가"]
        if not base_price:
            return None
        return round(close_now / base_price - 1, 4)

    rows = []
    for code in codes:
        if code not in today_df.index:
            continue
        close_now = today_df.loc[code, "종가"]

        pr1m = ret(m1_df, code, close_now)
        pr6m = ret(m6_df, code, close_now)
        pr_ytd = ret(ytd_df, code, close_now)
        pr1y = ret(y1_df, code, close_now)

        div_yield = None
        if not fund_df.empty and code in fund_df.index:
            div_val = fund_df.loc[code, "DIV"]
            div_yield = round(div_val / 100, 4) if pd.notna(div_val) else None

        market_cap = None
        if not cap_df.empty and code in cap_df.index:
            cap_val = cap_df.loc[code, "시가총액"]
            market_cap = round(cap_val / 1e8) if pd.notna(cap_val) else None

        rows.append({
            "종목코드": code,
            "종목명": name_map.get(code),
            "시가총액(억원)": market_cap,
            "현재가": close_now,
            "PR\n1개월": pr1m,
            "PR\n6개월": pr6m,
            "PR\nYTD": pr_ytd,
            "PR\n1년": pr1y,
            "최근 1년\n배당수익률": div_yield,
        })

    if progress_cb:
        progress_cb(1.0, "완료")
    out = pd.DataFrame(rows)[STOCK_COLS_KR]
    out = out.sort_values("시가총액(억원)", ascending=False, na_position="last").reset_index(drop=True)
    return out


# ----------------------------------------------------------------------------
# 4, 5, 6번 시트: 미국 (yfinance + 구성종목표)
# ----------------------------------------------------------------------------
_WEB_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
}


def _read_html_tables(url: str):
    """User-Agent 헤더 없이 pd.read_html(url)을 바로 호출하면 위키피디아 등에서
    403 Forbidden(봇 차단)이 발생할 수 있어, requests로 먼저 받아온 뒤 파싱한다."""
    import requests

    resp = requests.get(url, headers=_WEB_HEADERS, timeout=15)
    resp.raise_for_status()
    return pd.read_html(io.StringIO(resp.text))


def get_sp500_tickers() -> pd.DataFrame:
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    tables = _read_html_tables(url)
    df = tables[0][["Symbol", "Security"]].rename(
        columns={"Symbol": "종목코드", "Security": "종목명"}
    )
    df["종목코드"] = df["종목코드"].str.replace(".", "-", regex=False)
    return df


def get_nasdaq100_tickers() -> pd.DataFrame:
    # 주의: /wiki/Nasdaq-100 페이지에는 구성종목 표가 없다(지수 개요/역사 정보만 존재).
    # 실제 구성종목 리스트는 별도 페이지인 List_of_NASDAQ-100_companies 에 있다.
    url = "https://en.wikipedia.org/wiki/List_of_NASDAQ-100_companies"
    tables = _read_html_tables(url)
    for t in tables:
        cols = [str(c) for c in t.columns]
        if any("Ticker" in c for c in cols) or any("Symbol" in c for c in cols):
            ticker_col = [c for c in cols if "Ticker" in c or "Symbol" in c][0]
            name_col = [c for c in cols if "Company" in c or "Security" in c]
            name_col = name_col[0] if name_col else cols[0]
            df = t[[name_col, ticker_col]].rename(
                columns={name_col: "종목명", ticker_col: "종목코드"}
            )
            return df
    raise ValueError("나스닥100 표를 찾지 못했습니다.")


def _get_schd_nport_xml_url() -> str:
    """SCHWAB STRATEGIC TRUST(CIK 1454889)의 최신 NPORT-P 필링들 중
    SCHD 시리즈(S000034163)에 해당하는 필링의 primary_doc.xml URL을 찾는다."""
    import requests

    cik = "0001454889"  # SEC submissions API는 10자리 zero-padded CIK를 요구함
    cik_int = str(int(cik))  # Archives 경로는 앞의 0을 뗀 형태를 사용
    sec_headers = {"User-Agent": "ETF-List-App research-tool@example.com"}

    resp = requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json", headers=sec_headers, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    forms = data["filings"]["recent"]["form"]
    accessions = data["filings"]["recent"]["accessionNumber"]
    nport_filings = [a for f, a in zip(forms, accessions) if f == "NPORT-P"]

    for accession in nport_filings[:20]:  # 최근 20개 안에서 탐색 (분기별 필링이라 충분)
        accession_nodash = accession.replace("-", "")
        xml_url = f"https://www.sec.gov/Archives/edgar/data/{cik_int}/{accession_nodash}/primary_doc.xml"
        r = requests.get(xml_url, headers=sec_headers, timeout=15)
        if r.status_code == 200 and "S000034163" in r.text:  # SCHD 시리즈 ID
            return xml_url

    raise RuntimeError("SEC EDGAR에서 SCHD의 최신 N-PORT 필링을 찾지 못했습니다.")


def _cusip_to_ticker_batch(cusip_list, batch_size=10) -> dict:
    """OpenFIGI API(무료, 키 불필요)로 CUSIP 리스트를 티커로 일괄 변환."""
    import requests

    results = {}
    for i in range(0, len(cusip_list), batch_size):
        batch = cusip_list[i:i + batch_size]
        jobs = [{"idType": "ID_CUSIP", "idValue": c} for c in batch]
        resp = requests.post(
            "https://api.openfigi.com/v3/mapping",
            json=jobs,
            headers={"Content-Type": "application/json"},
            timeout=20,
        )
        if resp.status_code == 200:
            data = resp.json()
            for cusip, item in zip(batch, data):
                if "data" in item and item["data"]:
                    us_match = next((d for d in item["data"] if d.get("exchCode") == "US"), None)
                    chosen = us_match or item["data"][0]
                    results[cusip] = chosen.get("ticker")
                else:
                    results[cusip] = None
        else:
            for c in batch:
                results[c] = None
        time.sleep(0.3)  # 무료 사용 rate limit 방지
    return results


def get_schd_holdings() -> pd.DataFrame:
    """SCHD(Schwab US Dividend Equity ETF) 구성종목.

    SEC EDGAR의 공식 규제 공시(Form N-PORT)에서 전체 보유종목(CUSIP 기준)을 가져오고,
    OpenFIGI API로 CUSIP을 실제 매매 티커로 변환한다."""
    xml_url = _get_schd_nport_xml_url()

    import requests
    import xml.etree.ElementTree as ET

    sec_headers = {"User-Agent": "ETF-List-App research-tool@example.com"}
    r = requests.get(xml_url, headers=sec_headers, timeout=15)
    r.raise_for_status()
    root = ET.fromstring(r.text)

    holdings = []
    for elem in root.iter():
        tag = elem.tag.split("}")[-1]
        if tag == "invstOrSec":
            item = {}
            for child in elem:
                ctag = child.tag.split("}")[-1]
                if ctag in ("name", "cusip"):
                    item[ctag] = child.text
            holdings.append(item)

    cusips = [h.get("cusip") for h in holdings if h.get("cusip") and h.get("cusip") != "000000000"]
    ticker_map = _cusip_to_ticker_batch(cusips)

    rows = []
    for h in holdings:
        cusip = h.get("cusip")
        if not cusip or cusip == "000000000":
            continue
        ticker = ticker_map.get(cusip)
        if not ticker:
            continue
        rows.append({"종목코드": ticker, "종목명": h.get("name")})

    if not rows:
        raise RuntimeError("SCHD 구성종목을 SEC N-PORT에서 파싱했으나 결과가 비어 있습니다.")

    return pd.DataFrame(rows)


def calc_us_stock_metrics(ticker: str) -> dict:
    """yfinance로 미국 종목의 시가총액/주가/PR/배당수익률 계산.
    시가총액은 fast_info를 우선 사용(info보다 안정적, Colab에서 결측 9→0건으로 개선 확인),
    실패 시 info로 폴백한다."""
    import yfinance as yf

    try:
        t = yf.Ticker(ticker)
        hist = t.history(period="2y", auto_adjust=False)
        if hist.empty:
            return {}

        hist = hist.sort_index()
        last_close = hist["Close"].iloc[-1]
        last_date = hist.index[-1]

        def price_n_days_ago(days):
            target = last_date - pd.Timedelta(days=days)
            sub = hist[hist.index <= target]
            return sub["Close"].iloc[-1] if not sub.empty else None

        def price_at_ytd_start():
            ytd_start = pd.Timestamp(year=last_date.year, month=1, day=1, tz=hist.index.tz)
            sub = hist[hist.index <= ytd_start]
            if not sub.empty:
                return sub["Close"].iloc[-1]
            sub2 = hist[hist.index >= ytd_start]
            return sub2["Close"].iloc[0] if not sub2.empty else None

        p1m = price_n_days_ago(30)
        p6m = price_n_days_ago(182)
        p1y = price_n_days_ago(365)
        p_ytd = price_at_ytd_start()

        def pr(base):
            if base is None or base == 0:
                return None
            return round(last_close / base - 1, 4)

        div_yield = None
        try:
            divs = t.dividends
            if divs is not None and not divs.empty:
                one_year_ago = last_date - pd.Timedelta(days=365)
                recent_divs = divs[divs.index >= one_year_ago]
                div_yield = round(float(recent_divs.sum()) / last_close, 4)
        except Exception:
            pass

        market_cap = None
        try:
            fi = t.fast_info
            raw_cap = fi.get("marketCap") or fi.get("market_cap")
            if raw_cap:
                market_cap = round(raw_cap / 1e6)
        except Exception:
            pass
        if market_cap is None:
            try:
                info = t.info
                raw_cap = info.get("marketCap")
                market_cap = round(raw_cap / 1e6) if raw_cap else None
            except Exception:
                pass

        return {
            "시가총액(Mil)": market_cap,
            "현재가": round(float(last_close), 2),
            "PR\n1개월": pr(p1m),
            "PR\n6개월": pr(p6m),
            "PR\nYTD": pr(p_ytd),
            "PR\n1년": pr(p1y),
            "최근 1년\n배당수익률": div_yield,
        }
    except Exception:
        return {}


def process_us_stock_sheet(label, ticker_df_func, progress_cb=None, limit=None) -> pd.DataFrame:
    df = ticker_df_func()
    if limit:
        df = df.head(limit).reset_index(drop=True)
    rows = []
    n = len(df)
    for i, r in df.iterrows():
        metrics = calc_us_stock_metrics(r["종목코드"])
        rows.append({**r.to_dict(), **metrics})
        if progress_cb:
            progress_cb(min((i + 1) / n, 1.0), f"{label} {i + 1}/{n}")
        time.sleep(0.05)  # yfinance 과호출 방지
    out = pd.DataFrame(rows)
    for c in STOCK_COLS_US:
        if c not in out.columns:
            out[c] = None
    out = out[STOCK_COLS_US]
    out = out.sort_values("시가총액(Mil)", ascending=False, na_position="last").reset_index(drop=True)
    return out


# ----------------------------------------------------------------------------
# 엑셀 숫자 표기 프로토콜
#   - 비율(PR, 배당수익률 등): %, 소수점 2자리, 음수는 빨간색
#   - 일반 숫자(시가총액, 현재가 등): 1000단위 콤마, 음수는 빨간색
#   - 셀에 저장되는 실제 값은 그대로 숫자이며, 아래는 "표시 형식"만 지정한다
#     (정렬·필터·수식은 서식과 무관하게 정상 동작)
#
# 엑셀 레이아웃 프로토콜 (모든 시트 공통)
#   1행: 시트명과 동일한 제목 (굵게, 16pt)
#   2행: 빈 줄
#   3행: 헤더(컬럼명) - 굵게 + 자동 필터
#   4행부터: 데이터, 4행 기준으로 틀고정(제목/빈줄/헤더가 스크롤해도 고정)
#   열 너비: 각 컬럼 내용 길이에 맞춰 자동 조정
# ----------------------------------------------------------------------------
PERCENT_FORMAT = '0.00%;[Red]-0.00%'
NUMBER_FORMAT = '#,##0;[Red]-#,##0'

# 컬럼명에 아래 키워드가 포함되면 비율(%) 서식, 나머지 숫자형 컬럼은 콤마 서식으로 처리
PERCENT_KEYWORDS = ["TR", "PR", "배당", "분배율", "괴리율", "등락률"]
# 서식을 적용하지 않을 텍스트성 컬럼(그대로 둠)
TEXT_COLUMNS = {"종목코드", "ETF명", "종목명"}

HEADER_ROW = 3        # 컬럼명이 들어가는 행
DATA_START_ROW = 4    # 데이터가 시작되는 행


def _is_percent_column(col_name: str) -> bool:
    # "시가총액"처럼 순수 금액 컬럼은 비율이 아니므로 제외
    if "시가총액" in col_name:
        return False
    return any(kw in col_name for kw in PERCENT_KEYWORDS)


def write_formatted_sheet(writer, sheet_name: str, df: pd.DataFrame):
    """시트 하나를 제목/빈줄/헤더/데이터 구조로 작성하고,
    숫자 서식·헤더 굵게·자동 필터·틀고정·열 너비까지 한 번에 적용한다."""
    safe_name = sheet_name[:31]
    df.to_excel(writer, sheet_name=safe_name, index=False, header=False, startrow=DATA_START_ROW - 1)
    ws = writer.sheets[safe_name]

    n_cols = len(df.columns)

    # 1행: 제목 (시트명과 동일, 굵게 16pt)
    title_cell = ws.cell(row=1, column=1, value=sheet_name)
    title_cell.font = Font(bold=True, size=16)

    # 3행: 헤더(컬럼명), 굵게
    for col_idx, col_name in enumerate(df.columns, start=1):
        header_cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        header_cell.font = Font(bold=True)

    # 자동 필터: 헤더 행 전체 범위에 적용
    last_col_letter = get_column_letter(n_cols)
    last_row = DATA_START_ROW - 1 + len(df)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{last_row}"

    # 틀고정: 데이터 시작 행(4행) 위쪽 전체 고정
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # 숫자 서식(%, 콤마) 적용
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in TEXT_COLUMNS:
            continue
        if not pd.api.types.is_numeric_dtype(df[col_name]):
            continue
        col_letter = get_column_letter(col_idx)
        fmt = PERCENT_FORMAT if _is_percent_column(col_name) else NUMBER_FORMAT
        for row_idx in range(DATA_START_ROW, last_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = fmt

    # 열 너비: 헤더와 데이터 내용 중 가장 긴 길이에 맞춰 자동 조정
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        header_len = max((len(line) for line in str(col_name).split("\n")), default=0)
        try:
            data_len = df[col_name].astype(str).map(len).max()
        except Exception:
            data_len = 0
        if pd.isna(data_len):
            data_len = 0
        width = max(header_len, int(data_len)) + 4  # 여유폭
        ws.column_dimensions[col_letter].width = min(max(width, 8), 40)


# ----------------------------------------------------------------------------
# Streamlit UI
# ----------------------------------------------------------------------------
st.title("📊 ETF/종목 리스트 자동 업데이트 (pykrx + KIS API 버전)")
st.caption("국내 ETF / KOSPI200 / S&P500 / 나스닥100 / SCHD 구성종목")

st.markdown(
    """
**데이터 소스**
- 국내(1, 2번): `pykrx` (전종목 일괄조회, 종목별 반복호출 없음) + 국내 ETF 시가총액은 KIS Open API
- 미국(4, 5, 6번): `yfinance` + 위키피디아 / SEC EDGAR 구성종목표
- PR(주가수익률) 1개월/6개월/YTD/1년, 최근 1년 배당수익률을 각각 컬럼으로 제공
"""
)

login_ok = ensure_krx_login()
if not login_ok:
    st.warning("좌측 사이드바에 KRX 아이디/비밀번호를 입력해야 국내(1, 2번) 시트를 조회할 수 있습니다. "
               "(미국 시트만 쓸 경우엔 입력하지 않아도 됩니다.)")

kis_auth = ensure_kis_auth()
if not kis_auth:
    st.info("KIS App Key/Secret이 없어 국내 ETF 시가총액은 결측으로 처리됩니다. (나머지 항목은 정상)")

base_date = st.date_input("기준일", value=dt.date.today() - dt.timedelta(days=1))

sheet_options = st.multiselect(
    "업데이트할 시트 선택",
    ["1. 국내 상장 ETF", "2. KOSPI200종목",
     "4. 미국S&P500종목", "5. 미국나스닥100종목", "6. 미국 배당 ETF(SCHD)"],
    default=["1. 국내 상장 ETF", "2. KOSPI200종목", "4. 미국S&P500종목",
              "5. 미국나스닥100종목", "6. 미국 배당 ETF(SCHD)"],
)

run = st.button("🚀 업데이트 실행", type="primary")

if run:
    result_sheets = {}

    # ---------------- 국내 ETF ----------------
    if "1. 국내 상장 ETF" in sheet_options and not login_ok:
        st.error("KRX 로그인 정보가 없어 국내 상장 ETF 시트를 건너뜁니다. 사이드바에서 로그인해 주세요.")
    elif "1. 국내 상장 ETF" in sheet_options:
        st.info("국내 상장 ETF 전체 처리 중...")
        bar = st.progress(0.0)
        status = st.empty()
        try:
            def cb(p, msg):
                bar.progress(p)
                status.text(msg)
            out = build_domestic_etf_sheet(base_date, progress_cb=cb, kis_auth=kis_auth)
            result_sheets["국내 상장 ETF"] = out
            st.success(f"국내 ETF {len(out)}종목 완료")
        except Exception as e:
            st.error(f"국내 ETF 수집 실패: {e}")

    # ---------------- KOSPI200 ----------------
    if "2. KOSPI200종목" in sheet_options and not login_ok:
        st.error("KRX 로그인 정보가 없어 KOSPI200종목 시트를 건너뜁니다. 사이드바에서 로그인해 주세요.")
    elif "2. KOSPI200종목" in sheet_options:
        st.info("KOSPI200 종목 처리 중...")
        bar = st.progress(0.0)
        status = st.empty()
        try:
            def cb(p, msg):
                bar.progress(p)
                status.text(msg)
            out = build_kospi200_sheet(base_date, progress_cb=cb)
            result_sheets["KOSPI200종목"] = out
            st.success(f"KOSPI200 {len(out)}종목 완료")
        except Exception as e:
            st.error(f"KOSPI200 수집 실패: {e}")

    # ---------------- 미국 종목 시트 공통 ----------------
    def run_us_sheet(label, ticker_func, sheet_key):
        st.info(f"{label} 처리 중... (종목 수가 많아 몇 분 걸릴 수 있습니다)")
        bar = st.progress(0.0)
        status = st.empty()
        try:
            def cb(p, msg):
                bar.progress(p)
                status.text(msg)
            out = process_us_stock_sheet(label, ticker_func, progress_cb=cb)
            result_sheets[sheet_key] = out
            st.success(f"{label} {len(out)}종목 완료")
        except Exception as e:
            st.error(f"{label} 수집 실패: {e}")

    if "4. 미국S&P500종목" in sheet_options:
        run_us_sheet("S&P500", get_sp500_tickers, "미국S&P500종목")

    if "5. 미국나스닥100종목" in sheet_options:
        run_us_sheet("나스닥100", get_nasdaq100_tickers, "미국나스닥100종목")

    if "6. 미국 배당 ETF(SCHD)" in sheet_options:
        run_us_sheet("SCHD 구성종목", get_schd_holdings, "미국 배당 ETF(SCHD)")

    # ---------------- 결과 다운로드 ----------------
    if result_sheets:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for sheet_name, df in result_sheets.items():
                write_formatted_sheet(writer, sheet_name, df)
        buf.seek(0)

        st.download_button(
            "📥 결과 엑셀 다운로드",
            data=buf,
            file_name=f"ETF_종목_List_{base_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        for name, df in result_sheets.items():
            with st.expander(f"미리보기: {name} ({len(df)}행)"):
                st.dataframe(df.head(30))
